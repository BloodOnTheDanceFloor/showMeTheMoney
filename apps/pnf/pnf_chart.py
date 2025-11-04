import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import os
import math

class PNFChart:
    def __init__(self, excel_path, box_size_percent=0.02, reversal_boxes=1, box_size_value=None, box_size_choice=None):
        """
        初始化PNF图表
        
        参数:
        excel_path: Excel文件路径
        box_size_percent: 兼容旧参数（不再使用百分比计算）
        reversal_boxes: 反转所需的格子数，1表示1点图，3表示3点图
        """
        self.excel_path = excel_path
        self.box_size_percent = box_size_percent  # 保留以兼容旧调用，不参与计算
        self.reversal_boxes = reversal_boxes
        self.box_size_value = box_size_value  # 固定格值（绝对值），优先于百分比
        self.box_size_choice = box_size_choice  # 预设格值（从选项中选择）
        self.data = None
        self.box_size = None
        self.min_price = None
        self.max_price = None
        self.title = None  # Excel A1 标题
        self.suggested_box_size = None  # 基于股价中位数的建议格值
        self.chart_data = []  # 存储PNF图表数据
        self.mark_points = []  # 每个标记的元数据，用于交互显示
        
    def load_data(self):
        """加载Excel/TSV数据，并提取 A1 标题"""
        try:
            ext = os.path.splitext(self.excel_path)[1].lower()

            def _calc_box_size():
                # 计算格值（基于股价中位数的预设，或固定值）
                self.min_price = self.data['最低'].min()
                self.max_price = self.data['最高'].max()
                price_range = self.max_price - self.min_price
                # 计算股价中位数（使用收盘价）
                try:
                    median_price = float(self.data['收盘'].median())
                except Exception:
                    median_price = float(self.data[['开盘','收盘','最高','最低']].stack().median())

                # 基于中位数的建议格值
                def _suggest_by_median(m):
                    if m <= 5:
                        return 0.25
                    elif m <= 20:
                        return 0.5
                    elif m <= 100:
                        return 1.0
                    elif m <= 200:
                        return 2.0
                    elif m <= 500:
                        return 4.0
                    elif m <= 1000:
                        return 5.0
                    elif m <= 25000:
                        return 50.0
                    else:
                        return 500.0
                self.suggested_box_size = _suggest_by_median(median_price)

                if self.box_size_value is not None:
                    try:
                        fixed_box = float(self.box_size_value)
                    except (TypeError, ValueError):
                        print("固定格值无效：必须是大于0的浮点数")
                        return False
                    if fixed_box <= 0:
                        print("固定格值必须大于0")
                        return False
                    self.box_size = fixed_box
                elif self.box_size_choice is not None:
                    try:
                        choice_val = float(self.box_size_choice)
                        if choice_val <= 0:
                            raise ValueError("预设格值必须大于0")
                        self.box_size = choice_val
                    except Exception:
                        print("预设格值无效，改用建议值")
                        self.box_size = self.suggested_box_size
                else:
                    # 默认采用建议值
                    self.box_size = self.suggested_box_size
                return True

            def _validate_columns(df):
                # 去除列名空白并标准化类型
                try:
                    df.columns = df.columns.str.strip()
                except Exception:
                    df.columns = [str(c).strip() for c in df.columns]
                if '时间' in df.columns:
                    df['时间'] = df['时间'].astype(str).str.strip()
                required_columns = ['时间', '最高', '最低', '开盘', '收盘', '成交额']
                if not all(col in df.columns for col in required_columns):
                    # 若缺失必需列，尝试按位置映射修复
                    if df.shape[1] >= 6:
                        # 保存原始列名以便调试
                        original_columns = df.columns.tolist()
                        print(f"原始列名: {original_columns}")
                        
                        # 尝试多种常见的列顺序
                        column_orders = [
                            # 顺序1: 时间、最高、最低、开盘、收盘、成交额
                            ['时间', '最高', '最低', '开盘', '收盘', '成交额'],
                            # 顺序2: 时间、开盘、最高、最低、收盘、成交额
                            ['时间', '开盘', '最高', '最低', '收盘', '成交额'],
                            # 顺序3: 时间、收盘、最高、最低、开盘、成交额
                            ['时间', '收盘', '最高', '最低', '开盘', '成交额']
                        ]
                        
                        # 尝试每种列顺序
                        for i, column_order in enumerate(column_orders):
                            df_pos = df.iloc[:, :6].copy()
                            df_pos.columns = column_order
                            
                            # 标准化类型
                            df_pos['时间'] = df_pos['时间'].astype(str).str.strip()
                            try:
                                for c in ['最高', '最低', '开盘', '收盘', '成交额']:
                                    df_pos[c] = pd.to_numeric(df_pos[c], errors='coerce')
                                
                                # 检查数据是否合理（最高价应该 >= 最低价）
                                valid_rows = (df_pos['最高'] >= df_pos['最低']) & (df_pos['最高'] > 0) & (df_pos['最低'] > 0)
                                valid_percent = valid_rows.sum() / len(df_pos) if len(df_pos) > 0 else 0
                                
                                print(f"顺序{i+1}验证: 有效行比例 {valid_percent:.2%}")
                                
                                if valid_percent > 0.8:  # 80%以上的行数据合理
                                    self.data = df_pos.dropna(subset=['最高', '最低', '开盘', '收盘', '成交额']).reset_index(drop=True)
                                    print(f"使用位置映射修复列名 (顺序{i+1}: {'/'.join(column_order)})")
                                    return True
                            except Exception as e:
                                print(f"尝试顺序{i+1}时出错: {e}")
                                pass
                        
                        print("错误: 无法通过位置映射修复列名，数据格式不符合预期")
                        return False
                    else:
                        print("错误: Excel/TSV 文件缺少必要列，且列数不足以通过位置映射修复")
                        return False
                # 正常路径：强制数值化并去除缺失
                for c in ['最高', '最低', '开盘', '收盘', '成交额']:
                    df[c] = pd.to_numeric(df[c], errors='coerce')
                self.data = df.dropna(subset=['最高', '最低', '开盘', '收盘', '成交额']).reset_index(drop=True)
                return True

            if ext == '.xlsx':
                # 使用 openpyxl 读取标题与数据
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(filename=self.excel_path, read_only=True, data_only=True)
                    ws = wb.active
                    cell_val = ws.cell(row=1, column=1).value
                    if cell_val is not None:
                        self.title = str(cell_val).strip()
                except Exception as e:
                    print(f"读取A1标题失败: {e}")
                # 读取数据（跳过前2行，第三行为列名）
                self.data = pd.read_excel(self.excel_path, header=2, engine='openpyxl')
                if not _validate_columns(self.data):
                    return False
                if not _calc_box_size():
                    return False

            elif ext == '.xls':
                # 优先尝试 xlrd；如不可用则按 TSV 文本读取（许多‘.xls’实际为制表文本）
                used_tsv_fallback = False
                try:
                    import xlrd  # 可能未安装；如果缺失走回退
                    try:
                        book = xlrd.open_workbook(self.excel_path)
                        sheet = book.sheet_by_index(0)
                        try:
                            self.title = str(sheet.cell_value(0, 0)).strip()
                        except Exception:
                            pass
                    except Exception as e:
                        print(f"xlrd 打开 .xls 失败，改用TSV回退: {e}")
                        used_tsv_fallback = True
                except Exception:
                    used_tsv_fallback = True

                # 先尝试按 .xls 正常读取；若失败则回退 TSV
                if not used_tsv_fallback:
                    try:
                        self.data = pd.read_excel(self.excel_path, header=2, engine='xlrd')
                    except Exception as e:
                        print(f"pd.read_excel(xlrd) 读取 .xls 失败，改用TSV回退: {e}")
                        used_tsv_fallback = True

                if not used_tsv_fallback:
                    if not _validate_columns(self.data):
                        return False
                    if not _calc_box_size():
                        return False
                else:
                    # TSV 文本回退：优先 gbk，其次 utf-8
                    encoding_used = None
                    for enc in ('gbk', 'utf-8'):
                        try:
                            with open(self.excel_path, 'r', encoding=enc) as f:
                                first_line = f.readline()
                                if first_line:
                                    self.title = first_line.strip()
                            encoding_used = enc
                            break
                        except Exception:
                            continue
                    if encoding_used is None:
                        # 尝试使用 utf-8 非严格模式继续
                        try:
                            with open(self.excel_path, 'r', encoding='utf-8', errors='replace') as f:
                                first_line = f.readline()
                                if first_line:
                                    self.title = first_line.strip()
                            encoding_used = 'utf-8'
                        except Exception:
                            raise RuntimeError('无法识别 .xls 文本编码（尝试 gbk / utf-8 失败）')
                    # 读取数据：跳过前2行，第三行为表头
                    try:
                        self.data = pd.read_csv(self.excel_path, sep='\t', skiprows=2, encoding=encoding_used, engine='python')
                    except Exception:
                        # 若制表符失败，改用自动分隔符检测
                        self.data = pd.read_csv(self.excel_path, sep=None, skiprows=2, encoding=encoding_used, engine='python')
                    if not _validate_columns(self.data):
                        return False
                    if not _calc_box_size():
                        return False

            else:
                # 其他扩展名：尝试通用 TSV/CSV 读取，抓取第一行作为标题
                encoding_used = None
                for enc in ('gbk', 'utf-8'):
                    try:
                        with open(self.excel_path, 'r', encoding=enc) as f:
                            first_line = f.readline()
                            if first_line:
                                self.title = first_line.strip()
                        encoding_used = enc
                        break
                    except Exception:
                        continue
                if encoding_used is None:
                    # 无法读取标题也继续尝试 pandas 推断
                    encoding_used = 'utf-8'
                # 让 pandas 推断分隔符（可能性能较慢，但通用）
                try:
                    self.data = pd.read_csv(self.excel_path, sep=None, engine='python', skiprows=2, encoding=encoding_used)
                except Exception:
                    # 再次尝试制表分隔符
                    self.data = pd.read_csv(self.excel_path, sep='\t', skiprows=2, encoding=encoding_used)
                if not _validate_columns(self.data):
                    return False
                if not _calc_box_size():
                    return False

            print(f"数据加载成功: {len(self.data)}行")
            print(f"价格范围: {self.min_price:.2f} - {self.max_price:.2f}")
            print(f"格值: {self.box_size:.2f}")
            if self.title:
                print(f"A1 标题: {self.title}")
            return True
        except Exception as e:
            print(f"加载数据时出错: {e}")
            return False
    
    def calculate_pnf(self):
        """计算PNF图表数据"""
        if self.data is None or len(self.data) == 0:
            print("错误: 没有数据可处理")
            return False
        
        # 确保数据中包含所有必需的列
        required_columns = ['时间', '最高', '最低', '开盘', '收盘']
        if not all(col in self.data.columns for col in required_columns):
            print(f"错误: 数据缺少必需列。当前列: {self.data.columns.tolist()}")
            return False
            
        # 清空之前的图表数据
        self.chart_data = []
        
        # 初始化变量
        current_price = None  # 当前价格
        current_column = []   # 当前列
        current_direction = None  # 当前方向：'up' 或 'down'
        current_col_idx = 0   # 当前列的绝对索引
        self.mark_points = []
        
        # 处理第一天数据，确定初始方向
        try:
            first_day = self.data.iloc[0]
            high_price = first_day['最高']
            low_price = first_day['最低']
            
            # 确定初始方向（根据收盘价和开盘价）
            is_rising = first_day['收盘'] > first_day['开盘']
            current_direction = 'up' if is_rising else 'down'
        except Exception as e:
            print(f"处理第一天数据时出错: {e}")
            return False
        
        # 打印第一天数据和初始方向
        print(f"第一天日期: {first_day['时间']}")
        print(f"第一天开盘价: {first_day['开盘']}, 收盘价: {first_day['收盘']}")
        print(f"初始方向: {current_direction} (is_rising: {is_rising})")
        
        # 设置初始价格（对齐到栅格）
        def floor_to_box(p: float) -> float:
            return self.min_price + math.floor((p - self.min_price) / self.box_size) * self.box_size

        def ceil_to_box(p: float) -> float:
            return self.min_price + math.ceil((p - self.min_price) / self.box_size) * self.box_size

        # 辅助：记录一个标记及其元数据
        def record_mark(symbol: str, price: float, day_rec, col_idx: int):
            try:
                meta = {
                    'col': int(col_idx),
                    'price': float(price),
                    'symbol': symbol,
                    'date': day_rec['时间'],
                    'open': float(day_rec['开盘']),
                    'close': float(day_rec['收盘']),
                    'high': float(day_rec['最高']),
                    'low': float(day_rec['最低']),
                    'turnover': float(day_rec['成交额']),
                }
            except Exception:
                # 若某字段异常，尽可能保留基本信息
                meta = {
                    'col': int(col_idx),
                    'price': float(price),
                    'symbol': symbol,
                    'date': str(day_rec.get('时间', '')) if hasattr(day_rec, 'get') else str(day_rec),
                    'open': float(day_rec.get('开盘', price)) if hasattr(day_rec, 'get') else float(price),
                    'close': float(day_rec.get('收盘', price)) if hasattr(day_rec, 'get') else float(price),
                    'high': float(day_rec.get('最高', price)) if hasattr(day_rec, 'get') else float(price),
                    'low': float(day_rec.get('最低', price)) if hasattr(day_rec, 'get') else float(price),
                    'turnover': float(day_rec.get('成交额', 0)) if hasattr(day_rec, 'get') else 0,
                }
            self.mark_points.append(meta)

        if current_direction == 'up':
            # 上涨列：从当日低点的上方一个格开始，直到当日高点
            start_price = ceil_to_box(low_price)
            # 如果起点恰好不高于低点，则确保至少在低点之上一个格
            if start_price <= low_price:
                start_price = low_price + (self.box_size - ((low_price - self.min_price) % self.box_size))
                start_price = ceil_to_box(start_price)
            
            current_price = start_price
            while current_price <= high_price:
                current_column.append(('X', current_price))
                record_mark('X', current_price, first_day, current_col_idx)
                current_price += self.box_size
        else:
            # 下跌列：从当日高点的下方一个格开始，直到当日低点
            start_price = floor_to_box(high_price)
            # 如果起点恰好不低于高点，则确保至少在高点之下一个格
            if start_price >= high_price:
                start_price = high_price - ((high_price - self.min_price) % self.box_size)
                start_price = floor_to_box(start_price)
            
            current_price = start_price
            while current_price >= low_price:
                current_column.append(('O', current_price))
                record_mark('O', current_price, first_day, current_col_idx)
                current_price -= self.box_size
        
        # 如果第一列为空，进行兜底锚定，保证至少有一个点
        if len(current_column) == 0:
            if current_direction == 'up':
                anchor = ceil_to_box(low_price)
                if anchor > high_price:
                    anchor = floor_to_box(high_price)
                current_column.append(('X', anchor))
                record_mark('X', anchor, first_day, current_col_idx)
            else:
                anchor = floor_to_box(high_price)
                if anchor < low_price:
                    anchor = ceil_to_box(low_price)
                current_column.append(('O', anchor))
                record_mark('O', anchor, first_day, current_col_idx)
        
        # 添加第一列到图表
        self.chart_data.append((first_day['时间'], current_column))
        
        # 处理剩余的日期数据
        for i in range(1, len(self.data)):
            day = self.data.iloc[i]
            high_price = day['最高']
            low_price = day['最低']
        
            # 获取当前列的最后一个价格点
            last_price = current_column[-1][1]
        
            if current_direction == 'up':
                # 当前是上涨列
                
                # 检查是否可以继续上涨
                if high_price >= last_price + self.box_size:
                    # 可以继续在当前列添加X
                    temp_price = last_price + self.box_size
                    # 向上对齐到栅格
                    temp_price = ceil_to_box(temp_price)
                    while temp_price <= high_price:
                        current_column.append(('X', temp_price))
                        record_mark('X', temp_price, day, current_col_idx)
                        temp_price += self.box_size
                
                # 检查是否需要反转（下跌超过反转格数）
                elif low_price <= last_price - (self.reversal_boxes * self.box_size):
                    # 需要反转
                    if len(current_column) == 1:
                        # 如果当前列只有一个X：先补足反转所需 N 个格，再按实际跨越的额外格数补画
                        temp_price = last_price - self.box_size
                        # 先补足反转必需的 N 个格（N 点图）
                        for _ in range(self.reversal_boxes):
                            current_column.append(('O', temp_price))
                            record_mark('O', temp_price, day, current_col_idx)
                            temp_price -= self.box_size
                        # 计算额外跨越的格数（低点是否进一步跨越更多格）
                        extra_boxes = math.floor((last_price - low_price) / self.box_size) - self.reversal_boxes
                        while extra_boxes > 0:
                            current_column.append(('O', temp_price))
                            record_mark('O', temp_price, day, current_col_idx)
                            temp_price -= self.box_size
                            extra_boxes -= 1
                        current_direction = 'down'
                    else:
                        # 创建新的下跌列
                        new_column = []
                        new_col_idx = len(self.chart_data)
                        
                        # 从当前列的最后一个价格点开始向下，先满足反转格数
                        temp_price = last_price
                        boxes_to_add = self.reversal_boxes
                        while boxes_to_add > 0:
                            temp_price -= self.box_size
                            new_column.append(('O', temp_price))
                            record_mark('O', temp_price, day, new_col_idx)
                            boxes_to_add -= 1
                        
                        # 继续添加更多O直到达到最低价
                        while temp_price - self.box_size >= low_price:
                            temp_price -= self.box_size
                            new_column.append(('O', temp_price))
                            record_mark('O', temp_price, day, new_col_idx)
                        
                        # 添加新列到图表，并记录日期
                        self.chart_data.append((day['时间'], new_column))
                        
                        # 更新当前列和方向
                        current_column = new_column
                        current_col_idx = new_col_idx
                        current_direction = 'down'
            else:
                # 当前是下跌列
                
                # 检查是否可以继续下跌
                if low_price <= last_price - self.box_size:
                    # 可以继续在当前列添加O
                    temp_price = last_price - self.box_size
                    # 向下对齐到栅格
                    temp_price = floor_to_box(temp_price)
                    while temp_price >= low_price:
                        current_column.append(('O', temp_price))
                        record_mark('O', temp_price, day, current_col_idx)
                        temp_price -= self.box_size
                
                # 检查是否需要反转（上涨超过反转格数）
                elif high_price >= last_price + (self.reversal_boxes * self.box_size):
                    # 需要反转
                    if len(current_column) == 1:
                        # 如果当前列只有一个O：先补足反转所需 N 个格，再按实际跨越的额外格数补画
                        temp_price = last_price + self.box_size
                        # 先补足反转必需的 N 个格（N 点图）
                        for _ in range(self.reversal_boxes):
                            current_column.append(('X', temp_price))
                            record_mark('X', temp_price, day, current_col_idx)
                            temp_price += self.box_size
                        # 计算额外跨越的格数（高点是否进一步跨越更多格）
                        extra_boxes = math.floor((high_price - last_price) / self.box_size) - self.reversal_boxes
                        while extra_boxes > 0:
                            current_column.append(('X', temp_price))
                            record_mark('X', temp_price, day, current_col_idx)
                            temp_price += self.box_size
                            extra_boxes -= 1
                        current_direction = 'up'
                    else:
                        # 创建新的上涨列
                        new_column = []
                        new_col_idx = len(self.chart_data)
                        
                        # 从当前列的最后一个价格点开始向上，先满足反转格数
                        temp_price = last_price
                        boxes_to_add = self.reversal_boxes
                        while boxes_to_add > 0:
                            temp_price += self.box_size
                            new_column.append(('X', temp_price))
                            record_mark('X', temp_price, day, new_col_idx)
                            boxes_to_add -= 1
                        
                        # 继续添加更多X直到达到最高价
                        while temp_price + self.box_size <= high_price:
                            temp_price += self.box_size
                            new_column.append(('X', temp_price))
                            record_mark('X', temp_price, day, new_col_idx)
                        
                        # 添加新列到图表，并记录日期
                        self.chart_data.append((day['时间'], new_column))
                        
                        # 更新当前列和方向
                        current_column = new_column
                        current_col_idx = new_col_idx
                        current_direction = 'up'
        
        return True

    
    
    def plot_chart(self):
        """绘制PNF图表"""
        if not self.chart_data:
            print("错误: 没有PNF数据可绘制")
            return False
        
        # 创建图表
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # 设置Y轴范围（按格值对齐到网格）
        y_min = math.floor(self.min_price / self.box_size) * self.box_size
        y_max = math.ceil(self.max_price / self.box_size) * self.box_size
        ax.set_ylim(y_min, y_max)
        
        # 设置Y轴刻度为每一格值，并全部标出
        ticks = []
        t = y_min
        while t <= y_max + 1e-9:
            ticks.append(t)
            t += self.box_size
        ax.set_yticks(ticks)
        
        # 绘制水平网格线（与刻度一致）
        for ty in ticks:
            ax.axhline(y=ty, color='lightgray', linestyle='-', alpha=0.5)
        
        # 绘制PNF图表
        x_labels = []
        for col_idx, (date, column) in enumerate(self.chart_data):
            x_labels.append(date.strftime('%Y-%m-%d')) # 格式化日期作为X轴标签
            for symbol, price in column:
                if symbol == 'X':
                    ax.text(col_idx, price, 'X', ha='center', va='center', color='green', fontsize=12)
                else:  # symbol == 'O'
                    ax.text(col_idx, price, 'O', ha='center', va='center', color='red', fontsize=12)
        
        # 设置X轴刻度为列索引，并设置对应的日期标签
        ax.set_xticks(range(len(x_labels)))
        ax.set_xticklabels(x_labels, rotation=45, ha='right') # 旋转标签以避免重叠
        
        # 设置图表标题和标签
        reversal_type = f"{self.reversal_boxes}点图"
        if self.title:
            ax.set_title(self.title)
        else:
            ax.set_title(f"点数图 (PNF) - 格值: {self.box_size:.2f} - {reversal_type}")
        ax.set_xlabel("日期") # X轴标签改为日期
        ax.set_ylabel("价格")
        
        # 显示图表
        plt.tight_layout()
        plt.grid(True)
        
        # 保存图表
        output_path = os.path.join(os.path.dirname(self.excel_path), "pnf_chart.png")
        plt.savefig(output_path)
        print(f"PNF图表已保存至: {output_path}")
        
        plt.show()
        return True

def main():
    # 获取当前脚本所在目录
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, "工作簿1.xlsx")
    
    # 创建PNF图表对象
    pnf = PNFChart(excel_path, box_size_percent=0.02, reversal_boxes=1)
    
    # 加载数据
    if not pnf.load_data():
        return
    
    # 计算PNF图表数据
    if not pnf.calculate_pnf():
        return
    
    # 绘制图表
    pnf.plot_chart()

if __name__ == "__main__":
    main()